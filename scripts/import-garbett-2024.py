#!/usr/bin/env python
"""
import-garbett-2024.py — Migrate 180 2024 LFR records from Garbett Excel into ALHFRS schema.

Source: Copy of Live Facial Recognition Deployments.xlsx (Zoë Garbett / GLA)
Sheet:  "Deployment Data"
Rows:   2024 only (180 records)

Steps:
  1. Read Excel, filter year == 2024
  2. Map Garbett columns → ALHFRS schema
  3. Geocode via postcodes.io (free, no API key, UK postcodes)
  4. Deduplicate against existing met-police-lfr.json 2024 records (2 known)
  5. Write to staging file — do NOT touch met-police-lfr.json directly
  6. Print merge instructions

Usage:
  python import-garbett-2024.py
  python import-garbett-2024.py --start-id 367 --out staging/garbett-2024.json
  python import-garbett-2024.py --dry-run   # print first 5 records, no writes

After running:
  Review staging/garbett-2024.json, then manually merge into met-police-lfr.json
  or run: python merge-staging.py --staging staging/garbett-2024.json
"""

import argparse
import json
import sys
import time
import urllib.request
import urllib.error
from datetime import datetime
from pathlib import Path

# ── Paths ──────────────────────────────────────────────────────────────────────

EXCEL_PATH = Path(
    r"D:\Dev\jaredkrauss-art\Content & Assets\Featured Projects"
    r"\A London History of Facial Recognition Systems (ALHFRS)"
    r"\Assets\Research Material\Deployment Data"
    r"\Copy of Live Facial Recognition Deployments.xlsx"
)

DATASET_PATH = Path(r"D:\Dev\ALHFRS\data\deployments\met-police-lfr.json")
STAGING_DIR  = Path(r"D:\Dev\ALHFRS\data\staging")
DEFAULT_OUT  = STAGING_DIR / "garbett-2024.json"

# ── Column name mapping ────────────────────────────────────────────────────────
# Garbett Excel column → internal key.  Handles minor spelling variations.
# The script normalises column names (strip, lower) before matching.

COLUMN_MAP = {
    "deployment":           "location_name",
    "location":             "location_name",     # fallback spelling
    "date":                 "date_raw",
    "year":                 "year",
    "borough":              "borough",
    "ward (approx)":        "ward",
    "ward(approx)":         "ward",
    "ward":                 "ward",
    "ward code":            "ward_code",
    "post code (approx)":   "postcode",
    "postcode (approx)":    "postcode",
    "postcode":             "postcode",
    "faces seen":           "faces_scanned",
    "faces scanned":        "faces_scanned",
    "total alerts":         "total_alerts",
    "true alerts":          "true_alerts",
    "false alerts":         "false_alerts",
    "arrests":              "arrests",
    "watchlist size":       "watchlist_size",
    "duration":             "duration_hours",
    "min threshold setting":"threshold",
    "threshold":            "threshold",
    "lfr use case":         "use_case",
    "source":               "source_ref",
}

# ── Geocoding ──────────────────────────────────────────────────────────────────

_geocode_cache: dict[str, tuple[float, float]] = {}


def geocode_postcode(postcode: str) -> tuple[float, float] | None:
    """
    Look up lat/lon for a UK postcode via postcodes.io (free, no API key).
    Returns (lat, lon) or None if postcode is invalid/not found.
    Caches results to avoid duplicate calls.
    """
    if not postcode or not isinstance(postcode, str):
        return None

    pc = postcode.strip().upper().replace(" ", "")
    if not pc:
        return None

    if pc in _geocode_cache:
        return _geocode_cache[pc]

    try:
        url = f"https://api.postcodes.io/postcodes/{pc}"
        with urllib.request.urlopen(url, timeout=8) as resp:
            data = json.loads(resp.read())
        if data.get("status") == 200 and data.get("result"):
            lat = data["result"]["latitude"]
            lon = data["result"]["longitude"]
            _geocode_cache[pc] = (lat, lon)
            return lat, lon
    except Exception:
        pass

    # Try partial postcode (outward only — e.g. "SW1A" from "SW1A 1AA")
    outward = pc[:4].rstrip()
    if len(outward) >= 2 and outward != pc:
        try:
            url = f"https://api.postcodes.io/outcodes/{outward}"
            with urllib.request.urlopen(url, timeout=8) as resp:
                data = json.loads(resp.read())
            if data.get("status") == 200 and data.get("result"):
                lat = data["result"]["latitude"]
                lon = data["result"]["longitude"]
                _geocode_cache[pc] = (lat, lon)
                return lat, lon
        except Exception:
            pass

    _geocode_cache[pc] = None  # cache miss
    return None


# ── Date parsing ───────────────────────────────────────────────────────────────

def parse_date(raw) -> str | None:
    """Convert various date formats to YYYY-MM-DD."""
    if raw is None:
        return None
    if hasattr(raw, "strftime"):  # already a datetime/date object from openpyxl
        return raw.strftime("%Y-%m-%d")
    s = str(raw).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d %b %Y", "%d %B %Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


# ── Threshold rules ────────────────────────────────────────────────────────────

def threshold_for(date_str: str | None, garbett_threshold=None) -> float:
    """
    Prefer the Garbett-recorded threshold. Fall back to published MPS rules:
      Before 2024-07-11 → 0.60
      2024-07-11 to 2024-07-24 → 0.62
      2024-07-25+ → 0.64
    """
    if garbett_threshold and str(garbett_threshold).replace(".", "").isdigit():
        try:
            return float(garbett_threshold)
        except (ValueError, TypeError):
            pass
    if date_str:
        try:
            d = datetime.strptime(date_str, "%Y-%m-%d").date()
            from datetime import date as date_cls
            if d >= date_cls(2024, 7, 25):
                return 0.64
            if d >= date_cls(2024, 7, 11):
                return 0.62
        except ValueError:
            pass
    return 0.60


# ── Excel reading ──────────────────────────────────────────────────────────────

def read_garbett_excel(path: Path) -> list[dict]:
    """Read Deployment Data sheet, return list of normalised row dicts."""
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl not installed. Run:", file=sys.stderr)
        print("  pip install openpyxl --break-system-packages", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(str(path), data_only=True)

    # Find the right sheet
    sheet = None
    for name in wb.sheetnames:
        if "deployment" in name.lower():
            sheet = wb[name]
            break
    if sheet is None:
        sheet = wb.active
        print(f"  WARNING: No 'Deployment Data' sheet found, using active sheet: {sheet.title}", file=sys.stderr)

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Sheet is empty")

    # First row = headers
    raw_headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    headers = [h.lower() for h in raw_headers]

    print(f"  Sheet: {sheet.title}, {len(rows)-1} data rows, {len(headers)} columns", file=sys.stderr)
    print(f"  Columns: {raw_headers[:10]}{'...' if len(raw_headers) > 10 else ''}", file=sys.stderr)

    records = []
    for row in rows[1:]:
        d = {}
        for i, val in enumerate(row):
            if i < len(headers):
                col_key = COLUMN_MAP.get(headers[i], headers[i])
                d[col_key] = val
        records.append(d)

    return records


# ── Schema mapping ─────────────────────────────────────────────────────────────

def to_int(val) -> int | None:
    try:
        return int(float(val)) if val is not None and str(val).strip() not in ("", "None") else None
    except (ValueError, TypeError):
        return None


def to_float(val) -> float | None:
    try:
        return float(val) if val is not None and str(val).strip() not in ("", "None") else None
    except (ValueError, TypeError):
        return None


def map_to_alhfrs(row: dict, record_id: str, postcode_miss: list) -> dict | None:
    """Convert a Garbett row dict into ALHFRS schema. Returns None if row is unusable."""

    date_str = parse_date(row.get("date_raw"))
    if not date_str:
        return None

    location = str(row.get("location_name") or "").strip()
    borough  = str(row.get("borough") or "").strip()
    if not location and not borough:
        return None

    # Geocode
    postcode = str(row.get("postcode") or "").strip()
    coords = geocode_postcode(postcode) if postcode else None
    if coords:
        lat, lon = coords
    else:
        lat, lon = None, None
        if postcode:
            postcode_miss.append(postcode)

    threshold = threshold_for(date_str, row.get("threshold"))

    faces    = to_int(row.get("faces_scanned"))
    arrests  = to_int(row.get("arrests"))
    t_alerts = to_int(row.get("total_alerts"))
    tr_alerts = to_int(row.get("true_alerts"))
    fa_alerts = to_int(row.get("false_alerts"))

    # Infer false_alerts from total − true if not directly available
    if fa_alerts is None and t_alerts is not None and tr_alerts is not None:
        fa_alerts = t_alerts - tr_alerts

    # Build notes string
    notes_parts = []
    wl = to_int(row.get("watchlist_size"))
    if wl:
        notes_parts.append(f"Watchlist: {wl}")
    uc = str(row.get("use_case") or "").strip()
    if uc:
        notes_parts.append(f"Use case: {uc}")
    src = str(row.get("source_ref") or "").strip()
    if src:
        notes_parts.append(f"Source ref: {src}")

    record = {
        "id":                   record_id,
        "operator":             "Metropolitan Police",
        "operator_type":        "law_enforcement",
        "deployment_type":      "mobile",
        "vendor":               "NEC",
        "location_name":        location or borough,
        "borough":              borough or None,
        "ward":                 str(row.get("ward") or "").strip() or None,
        "ward_code":            str(row.get("ward_code") or "").strip() or None,
        "lat":                  lat,
        "lon":                  lon,
        "location_cluster_id":  None,
        "date_start":           date_str,
        "date_end":             date_str,
        "stated_purpose":       uc or None,
        "threshold":            threshold,
        "watchlist_size":       wl,
        "outcome_faces_scanned": faces,
        "outcome_alerts":       t_alerts,
        "outcome_true_alerts":  tr_alerts,
        "outcome_false_alerts": fa_alerts,
        "outcome_arrests":      arrests,
        "data_quality":         "confirmed",
        "source_url":           None,
        "source_type":          "garbett-gla-2024-deployment-grid",
        "notes":                ". ".join(notes_parts) if notes_parts else "",
    }

    return record


# ── Deduplication ──────────────────────────────────────────────────────────────

def load_existing_2024(dataset_path: Path) -> set[tuple[str, str]]:
    """Return set of (date_start, location_name_lower[:25]) for existing 2024 records."""
    try:
        data = json.loads(dataset_path.read_text(encoding="utf-8"))
        deps = data.get("deployments", data) if isinstance(data, dict) else data
        return {
            (r.get("date_start", ""), (r.get("location_name") or "").lower()[:25])
            for r in deps
            if (r.get("date_start") or "").startswith("2024")
        }
    except Exception as e:
        print(f"  WARNING: Could not load existing dataset: {e}", file=sys.stderr)
        return set()


def find_next_id(dataset_path: Path) -> int:
    """Return next available lfr-NNN integer."""
    try:
        data = json.loads(dataset_path.read_text(encoding="utf-8"))
        deps = data.get("deployments", data) if isinstance(data, dict) else data
        ids = []
        for r in deps:
            rid = str(r.get("id", ""))
            if rid.startswith("lfr-"):
                try:
                    ids.append(int(rid[4:]))
                except ValueError:
                    pass
        return max(ids) + 1 if ids else 400
    except Exception:
        return 400


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Migrate 2024 Garbett LFR records into ALHFRS staging JSON."
    )
    parser.add_argument("--start-id",  type=int, default=None,
                        help="First lfr-NNN ID to assign (auto-detected from dataset if omitted)")
    parser.add_argument("--out",       default=str(DEFAULT_OUT),
                        help=f"Staging output path (default: {DEFAULT_OUT})")
    parser.add_argument("--year",      type=int, default=2024,
                        help="Year to filter (default: 2024)")
    parser.add_argument("--dry-run",   action="store_true",
                        help="Print first 5 mapped records and stats, no file writes")
    parser.add_argument("--no-geocode", action="store_true",
                        help="Skip postcodes.io lookups (lat/lon will be null)")
    args = parser.parse_args()

    print(f"\n=== Garbett → ALHFRS import (year={args.year}) ===", file=sys.stderr)

    # ── Read Excel ──
    if not EXCEL_PATH.exists():
        print(f"ERROR: Excel file not found: {EXCEL_PATH}", file=sys.stderr)
        sys.exit(1)

    all_rows = read_garbett_excel(EXCEL_PATH)

    # Filter to target year
    year_rows = []
    for row in all_rows:
        y = row.get("year")
        try:
            if y is not None and int(float(str(y))) == args.year:
                year_rows.append(row)
        except (ValueError, TypeError):
            pass

    print(f"  {len(all_rows)} total rows → {len(year_rows)} in {args.year}", file=sys.stderr)

    # ── ID assignment ──
    start_id = args.start_id or find_next_id(DATASET_PATH)
    print(f"  Starting at ID: lfr-{start_id:03d}", file=sys.stderr)

    # ── Existing 2024 records (to skip duplicates) ──
    existing_2024 = load_existing_2024(DATASET_PATH)
    print(f"  Existing 2024 records in dataset: {len(existing_2024)}", file=sys.stderr)

    # ── Map records ──
    records = []
    skipped_invalid = 0
    skipped_dup = 0
    postcode_misses = []
    counter = start_id

    for row in year_rows:
        rec_id = f"lfr-{counter:03d}"
        if args.no_geocode:
            row["postcode"] = None  # suppress geocoding

        rec = map_to_alhfrs(row, rec_id, postcode_misses)
        if rec is None:
            skipped_invalid += 1
            continue

        # Dedup check
        key = (rec["date_start"], (rec["location_name"] or "").lower()[:25])
        if key in existing_2024:
            print(f"  DEDUP: skip {rec['date_start']} {rec['location_name'][:30]}", file=sys.stderr)
            skipped_dup += 1
            continue

        records.append(rec)
        existing_2024.add(key)  # prevent self-duplication within batch
        counter += 1

        # Rate limit geocoding to avoid hammering postcodes.io
        if not args.no_geocode and len(records) % 20 == 0:
            time.sleep(0.5)

    print(f"\n  Results:", file=sys.stderr)
    print(f"    Mapped:          {len(records)}", file=sys.stderr)
    print(f"    Skipped invalid: {skipped_invalid}", file=sys.stderr)
    print(f"    Skipped dedup:   {skipped_dup}", file=sys.stderr)
    print(f"    Postcode misses: {len(postcode_misses)}", file=sys.stderr)
    if postcode_misses:
        print(f"    Missing postcodes: {postcode_misses[:10]}", file=sys.stderr)

    geocoded = sum(1 for r in records if r.get("lat") is not None)
    print(f"    Geocoded:        {geocoded}/{len(records)}", file=sys.stderr)

    if args.dry_run:
        print("\n=== DRY RUN — first 5 records ===")
        for r in records[:5]:
            print(json.dumps(r, indent=2, ensure_ascii=False))
        print(f"\n[Would write {len(records)} records to {args.out}]")
        return

    # ── Write staging output ──
    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    output = {
        "import_date":    datetime.now().strftime("%Y-%m-%d"),
        "source":         "Garbett GLA 2024 LFR deployment grid",
        "year_filtered":  args.year,
        "record_count":   len(records),
        "id_range":       f"lfr-{start_id:03d} to lfr-{counter-1:03d}",
        "geocoded":       geocoded,
        "postcode_misses": postcode_misses,
        "deployments":    records,
    }

    out_path.write_text(json.dumps(output, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"\n  ✓ Written {len(records)} records to {out_path}", file=sys.stderr)
    print(f"  Next available ID: lfr-{counter:03d}", file=sys.stderr)
    print(f"\n  Review staging file, then merge with:", file=sys.stderr)
    print(f"    python merge-staging.py --staging \"{out_path}\"", file=sys.stderr)


if __name__ == "__main__":
    main()
